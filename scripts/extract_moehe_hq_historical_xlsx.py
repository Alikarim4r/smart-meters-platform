#!/usr/bin/env python3
"""Extract MOEHE HQ historical readings from the Feb 2020–May 2026 XLSX.

Supports two sheet layouts:
  - legacy (Feb–Oct 2020): dates as rows, meters as columns
  - modern (Nov 2020–May 2026): meters as rows, dates as columns

Usage:
  .venv-import/bin/python scripts/extract_moehe_hq_historical_xlsx.py
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
IMPORT_DIR = ROOT / "imports" / "moehe_hq_reports"
DEFAULT_XLSX = IMPORT_DIR / "utility_readings_feb2020_may2026.xlsx"
DEFAULT_METERS = IMPORT_DIR / "moehe_hq_meters_from_reports.csv"
DEFAULT_OUT = IMPORT_DIR / "moehe_hq_readings_feb2020_may2026.csv"
DEFAULT_SUMMARY = IMPORT_DIR / "moehe_hq_historical_extract_summary.json"

SITE_NAME = "MOEHE HQ / Permanent HQ"
SOURCE_FILE = "utility_readings_feb2020_may2026.xlsx"
EXCEL_EPOCH = date(1899, 12, 30)

MONTH_SHEET_RE = re.compile(
    r"^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{2})$",
    re.I,
)
MONTH_NUM = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}

SECTION_HEADERS = {
    "utility readings",
    "kahramaa electrical meter number",
    "domestic water meter no. / location",
    "tse water meter no. / location",
    "diesel meter no. / location",
    "energy & flow meters / location",
    "total",
    "readings (m3)",
    "readings (liters)",
    "consumption reading",
    "consumption readings",
}


def normalize_label(value: Any) -> str:
    text = str(value or "")
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def slug_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def excel_serial_to_date(value: float | int) -> date | None:
    try:
        serial = float(value)
    except (TypeError, ValueError):
        return None
    if serial < 30000 or serial > 60000:
        return None
    return EXCEL_EPOCH + timedelta(days=int(serial))


def parse_date_cell(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return excel_serial_to_date(value)
    text = normalize_label(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return None


def is_number(value: Any) -> bool:
    if value is None or value == "":
        return False
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return True
    text = normalize_label(value).replace(",", "")
    if text.lower() in {"n/a", "-", "na", "none", "null"}:
        return False
    try:
        float(text)
        return True
    except ValueError:
        return False


def as_float(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    return float(normalize_label(value).replace(",", ""))


def load_meter_catalog(path: Path) -> tuple[list[dict[str, str]], dict[str, dict[str, str]]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.DictReader(handle))

    aliases: dict[str, dict[str, str]] = {}
    for row in rows:
        code = row["meter_code"].strip()
        candidates = {
            code,
            row.get("name_en", ""),
            row.get("raw_label", ""),
            f"{code} {row.get('name_en', '')}",
        }
        # Common aliases from early sheets.
        name = row.get("name_en", "").strip()
        if name:
            candidates.add(name)
            candidates.add(name.replace(" ", ""))
        if code.startswith("125635"):
            candidates.add(name)  # LVP-1 etc.

        for candidate in candidates:
            key = slug_key(candidate)
            if key:
                aliases[key] = row

        # Prefix match helpers for labels like "1219053 (Beside Entrance..."
        aliases[slug_key(code)] = row

    # Explicit legacy aliases.
    extra = {
        "lvp1": "1256358",
        "lvp2a": "1256359",
        "lvp3a": "1256360",
        "lvp4": "1256361",
        "lvp5": "1256362",
        "lvp6": "1256363",
        "lvp7a": "1256364",
        "lvp8a": "1256365",
        "tsewater": "KF-18540133",  # best early single TSE column → main TSE tank
        "chwloop1": "CHW-LOOP-1",
        "chwloop2": "CHW-LOOP-2",
        "chwloop3": "CHW-LOOP-3",
        "makeupkahramma": "MAKEUP-KAHRAMMA",
        "makeupro": "MAKEUP-RO",
        "stormwater": "STORM-WATER",
        "colingtowerblowdown": "COLING-TOWER--BLOWDOWN",
        "coolingtowerblowdown": "COLING-TOWER--BLOWDOWN",
        "etd1000": "ETD-1000",
        "tms2000": "TMS-2000",
        "cap3000": "CAP-3000",
        "khawatermeterforb1": "KHA-WATER-METER-FOR-B1",
        "khawatermeterforb2": "KHA-WATER-METER-FOR-B2",
        "khawatermeterforb3": "KHA-WATER-METER-FOR-B3",
        "khawatermeterforb4": "KHA-WATER-METER-FOR-B4",
        "khawatermeterforb5": "KHA-WATER-METER-FOR-B5",
        "khawatermeterforgr1": "KHA-WATER-METER-FOR-GR1",
        "khawatermeterforgr2": "KHA-WATER-METER-FOR-GR2",
        "19aci005333": "19ACI-005333",
    }
    by_code = {row["meter_code"]: row for row in rows}
    for alias, code in extra.items():
        if code in by_code:
            aliases[alias] = by_code[code]

    return rows, aliases


def resolve_meter(
    label: str,
    aliases: dict[str, dict[str, str]],
) -> dict[str, str] | None:
    text = normalize_label(label)
    if not text:
        return None
    lowered = text.lower()
    if lowered in SECTION_HEADERS or lowered.startswith("note:"):
        return None
    if "total" == lowered or lowered.startswith("total "):
        return None

    key = slug_key(text)
    if key in aliases:
        return aliases[key]

    # Leading meter code / id before spaces or parentheses.
    m = re.match(r"^([A-Za-z0-9\-]+)", text)
    if m:
        code_key = slug_key(m.group(1))
        if code_key in aliases:
            return aliases[code_key]

    # Contains known code as prefix of raw_label.
    for alias_key, row in aliases.items():
        code = slug_key(row["meter_code"])
        if code and key.startswith(code):
            return row
        raw = slug_key(row.get("raw_label", ""))
        if raw and (key.startswith(raw[:12]) or raw.startswith(key[:12])):
            return row
        name = slug_key(row.get("name_en", ""))
        if name and (name == key or key.startswith(name) or name.startswith(key)):
            return row

    return None


def category_for_section(section: str) -> tuple[str, str]:
    s = section.lower()
    if "electrical" in s or "kahramaa electrical" in s:
        return "electricity", "kwh"
    if "diesel" in s:
        return "fuel", "liter"
    if "energy" in s or "flow" in s:
        # Default for non-CHW rows in Energy & Flow section (makeup, storm, etc.)
        return "flow", "m3"
    return "water", "m3"


def sheet_is_modern(rows: list[tuple[Any, ...]]) -> bool:
    sample = " | ".join(
        normalize_label(c) for row in rows[:5] for c in row[:8] if c is not None
    )
    return "Utility Readings" in sample or "Kahramaa" in sample


def extract_modern_sheet(
    sheet_name: str,
    rows: list[tuple[Any, ...]],
    aliases: dict[str, dict[str, str]],
    unmatched: Counter[str],
) -> list[dict[str, str]]:
    if not rows:
        return []

    # Row 2 (index 1) typically holds Excel serial dates starting at col B.
    date_row = rows[1] if len(rows) > 1 else ()
    dates: dict[int, date] = {}
    for col_idx, value in enumerate(date_row):
        if col_idx == 0:
            continue
        parsed = parse_date_cell(value)
        if parsed:
            dates[col_idx] = parsed

    if not dates:
        # Fallback: scan first 5 rows for a date row.
        for row in rows[:5]:
            found: dict[int, date] = {}
            for col_idx, value in enumerate(row):
                if col_idx == 0:
                    continue
                parsed = parse_date_cell(value)
                if parsed:
                    found[col_idx] = parsed
            if len(found) >= 5:
                dates = found
                break

    section = "unknown"
    out: list[dict[str, str]] = []
    for row_idx, row in enumerate(rows, start=1):
        if not row:
            continue
        label = normalize_label(row[0])
        if not label:
            continue
        lower = label.lower()
        if lower in SECTION_HEADERS or "meter no" in lower or "meters / location" in lower:
            if "electrical" in lower:
                section = "electrical"
            elif "domestic" in lower:
                section = "domestic"
            elif "tse" in lower:
                section = "tse"
            elif "diesel" in lower:
                section = "diesel"
            elif "energy" in lower or "flow" in lower:
                section = "flow"
            continue
        if lower == "total" or lower.startswith("total"):
            continue

        meter = resolve_meter(label, aliases)
        if not meter:
            # Only count labels that look like meter rows (have some numeric cells).
            if any(is_number(v) for v in row[1:]):
                unmatched[label] += 1
            continue

        cat_code = meter["category_code"].strip().lower()
        unit_code = meter["unit_code"].strip().lower()
        # CHW loops are cooling energy (GJ), never water/flow fallback.
        if meter["meter_code"].strip().upper() in {
            "CHW-LOOP-1",
            "CHW-LOOP-2",
            "CHW-LOOP-3",
        }:
            cat_code = "btu"
            unit_code = "gj"
        elif cat_code == "flow":
            # Keep flow for importer Option B (non-CHW energy/flow rows).
            pass

        for col_idx, value in enumerate(row):
            if col_idx == 0 or col_idx not in dates:
                continue
            if not is_number(value):
                continue
            reading_date = dates[col_idx]
            out.append(
                {
                    "site_name": SITE_NAME,
                    "meter_code": meter["meter_code"],
                    "reading_date": reading_date.isoformat(),
                    "raw_value": f"{as_float(value):.6g}",
                    "unit_code": unit_code,
                    "category_code": cat_code,
                    "source_file": SOURCE_FILE,
                    "source_sheet": sheet_name,
                    "source_row": str(row_idx),
                    "note": (
                        f"Imported from historical utility workbook "
                        f"({sheet_name}, modern layout)"
                    ),
                }
            )
    return out


def extract_legacy_sheet(
    sheet_name: str,
    rows: list[tuple[Any, ...]],
    aliases: dict[str, dict[str, str]],
    unmatched: Counter[str],
) -> list[dict[str, str]]:
    """Early sheets: row 2 headers LVP-1.. / TSE / Diesel; dates down column A."""
    if len(rows) < 4:
        return []

    header = rows[1]
    meter_cols: dict[int, dict[str, str]] = {}
    for col_idx, value in enumerate(header):
        if col_idx == 0:
            continue
        label = normalize_label(value)
        if not label or label.lower() == "date" or label.lower() == "total":
            continue
        if label.lower() == "diesel":
            # Aggregate diesel column — cannot split across CAP/ETD/TMS.
            unmatched[f"LEGACY_AGGREGATE:{label}"] += 1
            continue
        meter = resolve_meter(label, aliases)
        if not meter:
            unmatched[label] += 1
            continue
        meter_cols[col_idx] = meter

    out: list[dict[str, str]] = []
    consumption_section = False
    for row_idx, row in enumerate(rows, start=1):
        if not row:
            continue
        first = normalize_label(row[0])
        if "consumption per day" in first.lower():
            consumption_section = True
            continue
        if consumption_section:
            # Skip derived consumption block (not cumulative readings).
            continue
        if first.lower() in {"date", ""} and row_idx <= 3:
            continue

        reading_date = parse_date_cell(row[0])
        if not reading_date:
            continue

        for col_idx, meter in meter_cols.items():
            if col_idx >= len(row):
                continue
            value = row[col_idx]
            if not is_number(value):
                continue
            out.append(
                {
                    "site_name": SITE_NAME,
                    "meter_code": meter["meter_code"],
                    "reading_date": reading_date.isoformat(),
                    "raw_value": f"{as_float(value):.6g}",
                    "unit_code": meter["unit_code"].strip().lower(),
                    "category_code": meter["category_code"].strip().lower(),
                    "source_file": SOURCE_FILE,
                    "source_sheet": sheet_name,
                    "source_row": str(row_idx),
                    "note": (
                        f"Imported from historical utility workbook "
                        f"({sheet_name}, legacy layout)"
                    ),
                }
            )
    return out


def load_sheet_rows(ws: Any) -> list[tuple[Any, ...]]:
    return [tuple(row) for row in ws.iter_rows(values_only=True)]


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--meters-csv", type=Path, default=DEFAULT_METERS)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    parser.add_argument("--summary", type=Path, default=DEFAULT_SUMMARY)
    args = parser.parse_args()

    if not args.xlsx.exists():
        raise SystemExit(f"Missing workbook: {args.xlsx}")
    if not args.meters_csv.exists():
        raise SystemExit(f"Missing meters CSV: {args.meters_csv}")

    meters, aliases = load_meter_catalog(args.meters_csv)
    wb = load_workbook(args.xlsx, data_only=True, read_only=True)

    all_rows: list[dict[str, str]] = []
    unmatched: Counter[str] = Counter()
    per_sheet: dict[str, int] = {}
    layout_counts = Counter()

    for sheet_name in wb.sheetnames:
        if not MONTH_SHEET_RE.match(sheet_name):
            continue
        ws = wb[sheet_name]
        rows = load_sheet_rows(ws)
        if sheet_is_modern(rows):
            layout_counts["modern"] += 1
            extracted = extract_modern_sheet(sheet_name, rows, aliases, unmatched)
        else:
            layout_counts["legacy"] += 1
            extracted = extract_legacy_sheet(sheet_name, rows, aliases, unmatched)
        per_sheet[sheet_name] = len(extracted)
        all_rows.extend(extracted)

    wb.close()

    # Deduplicate by meter_code + reading_date (keep last = later sheet wins).
    dedup: dict[tuple[str, str], dict[str, str]] = {}
    for row in all_rows:
        dedup[(row["meter_code"], row["reading_date"])] = row
    final_rows = sorted(
        dedup.values(),
        key=lambda r: (r["reading_date"], r["meter_code"]),
    )

    fieldnames = [
        "site_name",
        "meter_code",
        "reading_date",
        "raw_value",
        "unit_code",
        "category_code",
        "source_file",
        "source_sheet",
        "source_row",
        "note",
    ]
    args.out.parent.mkdir(parents=True, exist_ok=True)
    with args.out.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(final_rows)

    by_meter = Counter(r["meter_code"] for r in final_rows)
    by_month = Counter(r["reading_date"][:7] for r in final_rows)
    summary = {
        "source_xlsx": str(args.xlsx),
        "meters_catalog": len(meters),
        "sheets_processed": len(per_sheet),
        "layout_counts": dict(layout_counts),
        "readings_raw": len(all_rows),
        "readings_deduped": len(final_rows),
        "date_min": final_rows[0]["reading_date"] if final_rows else None,
        "date_max": final_rows[-1]["reading_date"] if final_rows else None,
        "per_sheet": per_sheet,
        "per_meter": dict(by_meter),
        "per_month": dict(sorted(by_month.items())),
        "unmatched_labels": unmatched.most_common(50),
        "output_csv": str(args.out),
    }
    args.summary.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    print(f"Wrote {len(final_rows)} readings → {args.out}")
    print(f"Summary → {args.summary}")
    print(f"Layouts: {dict(layout_counts)}")
    print(f"Date range: {summary['date_min']} … {summary['date_max']}")
    if unmatched:
        print("Top unmatched labels:")
        for label, count in unmatched.most_common(15):
            print(f"  {count:5d}  {label}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
